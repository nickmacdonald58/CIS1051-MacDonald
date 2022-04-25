#Phillies Results Generator
import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, Series

def resultGrab():
   wbO= load_workbook('Optimization w Results.xlsm')
   wsO=wbO['Reader']
   wbR= load_workbook('Results.xlsx')
   wbR.create_sheet('Results',0)
   wsR=wbR['Results']
   wsR['A1']='Game Number'
   wsR['B1']='Player Name'
   wsR['C1']='Rest Days'
   wsR['D1']='Run Share'
   wsR['E1']='Expected IP'
   i=2

   for row in range(2,3889):  #3889 rows ##Ignores 0 selections
      if float(wsO.cell(row=row,column=12).value)>.9:
         wsR.cell(row=i,column=1,value=wsO.cell(row=row,column=1).value)
         wsR.cell(row=i,column=2,value=wsO.cell(row=row,column=4).value)
         wsR.cell(row=i,column=3,value=wsO.cell(row=row,column=15).value)
         wsR.cell(row=i,column=4,value=wsO.cell(row=row,column=13).value)
         wsR.cell(row=i,column=5,value=wsO.cell(row=row,column=14).value)
         i=i+1
      

   wbR.save('Results.xlsx')
   wbO.close()
   wbR.close()

def resultSummary():
   resultGrab()
   
   gameDict={}
   starter=[]
   bullpen=[]
   wbR=load_workbook('Results.xlsx')
   wbR.create_sheet('Summary',1)
   wsS=wbR['Summary']
   wsR=wbR['Results']
   wsS['A1']='Player Name'
   wsS['B1']='Appearances'
   wsS['C1']='S/R'
   i=2

   for row in range(2,len(wsR['B'])+1):   #Creates Dictionary
      if wsR.cell(row=row,column=2).value in gameDict:
         gameDict[wsR.cell(row=row,column=2).value]=gameDict[wsR.cell(row=row,column=2).value]+1
      else:
         gameDict[wsR.cell(row=row,column=2).value]=1

   starter.append(wsR.cell(row=2,column=1).value)   #Creates list of starters and bullpen
   for row in range(3,len(wsR['B'])):
      if int(wsR.cell(row=row,column=1).value) != int(wsR.cell(row=row-1,column=1).value) and wsR.cell(row=row,column=1).value not in starter:
         starter.append(wsR.cell(row=row,column=2).value)
      if int(wsR.cell(row=row,column=1).value) == int(wsR.cell(row=row-1,column=1).value) and wsR.cell(row=row,column=1).value not in bullpen:
         bullpen.append(wsR.cell(row=row,column=2).value)
      
   
   for key in gameDict:   #Formulates first table in summary sheet
      wsS.cell(row=i,column=1,value=key)
      wsS.cell(row=i,column=2,value=gameDict[key])
      if key in starter:
         wsS.cell(row=i,column=3,value='starter')
      else:
         wsS.cell(row=i,column=3,value='bullpen')
      i=i+1

   wsS['E1']='Starter Name'   #Creates starter and bullpen tables
   wsS['F1']='Appearances'
   wsS['H1']='Bullpen Name'
   wsS['I1']='Appearances'
   j=2
   k=2
   
   for row in range(2,len(wsS['B'])+1):   
      if wsS.cell(row=row,column=3).value=='starter':
         wsS.cell(row=j,column=5,value=wsS.cell(row=row,column=1).value)
         wsS.cell(row=j,column=6,value=wsS.cell(row=row,column=2).value)
         j=j+1
      elif wsS.cell(row=row,column=3).value=='bullpen':
         wsS.cell(row=k,column=8,value=wsS.cell(row=row,column=1).value)
         wsS.cell(row=k,column=9,value=wsS.cell(row=row,column=2).value)
         k=k+1
      
   
   wbR.save('Results.xlsx')
   
def resultChart():
   resultSummary()
   wbR=load_workbook('Results.xlsx')
   wbR.create_sheet('Chart',2)
   wsC=wbR['Chart']
   wsS=wbR['Summary']

   #Chart for starters
   valuesS=Reference(wsS,min_col=6,min_row=2,max_col=6,max_row=8)
   categoriesS=Reference(wsS,min_col=5,min_row=2,max_col=5,max_row=8)
   chartS=BarChart()
   chartS.title='Starter Appearances'
   chartS.x_axis.title='Player Name'
   chartS.y_axis.title='Appearances'
   chartS.add_data(valuesS)
   chartS.set_categories(categoriesS)
   wsC.add_chart(chartS,'B2')

   #Chart for bullpen
   valuesB=Reference(wsS,min_col=9,min_row=2,max_col=9,max_row=18)
   categoriesB=Reference(wsS,min_col=8,min_row=2,max_col=8,max_row=18)
   chartB=BarChart()
   chartB.title='Bullpen Appearances'
   chartB.x_axis.title='Player Name'
   chartB.y_axis.title='Appearances'
   chartB.add_data(valuesB)
   chartB.set_categories(categoriesB)
   wsC.add_chart(chartB,'K2')

   wbR.save('Results.xlsx')  

resultChart()




            
